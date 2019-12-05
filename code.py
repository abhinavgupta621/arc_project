import xlrd
import os
cwd=(os.getcwd())+'\\data'
loc = (cwd)+'\\courses.xls'
loc2= (cwd)+'\\data1.xls'
wb = xlrd.open_workbook(loc, formatting_info=True)
wb2 = xlrd.open_workbook(loc2)
sheet1=wb.sheet_by_index(1)
sheet0=wb.sheet_by_index(0)
sheet2=wb2.sheet_by_index(0)




for i in range(2):
	if(sheet1.cell_value(0, i).lower()=='discipline'):
		discipline=i
	else:
		code=i
		
Dict={}
for i in range(1,sheet1.nrows):
	Dict[sheet1.cell_value(i,code)[0:4].upper()]=sheet1.cell_value(i, discipline)

for i in range(sheet0.ncols):
	if(sheet0.cell_value(0, i).lower()=='course code'):
		course_code=i
	elif(sheet0.cell_value(0, i).lower()=='tag'):
		tag=i
	elif(sheet0.cell_value(0, i).lower()=='course name'):
		course_name=i
	elif(sheet0.cell_value(0, i).lower()=='units'):
		units=i

course_H101={}
elec_H101={}
course_H103={}
elec_H103={}
course_H106={}
elec_H106={}
course_H123={}
elec_H123={}
course_H129={}
elec_H129={}
course_H140={}
elec_H140={}
course_H141={}
elec_H141={}
course_H151={}
elec_H151={}
course_H152={}
elec_H152={}
unit_H101={}
unit_H103={}
unit_H106={}
unit_H123={}
unit_H129={}
unit_H140={}
unit_H141={}
unit_H151={}
unit_H152={}
blue_check={}
for i in range(sheet0.nrows):
	xfx = sheet0.cell_xf_index(i, tag)
	xf = wb.xf_list[xfx]
	bgx = xf.background.pattern_colour_index
	if(bgx == 40):
		blue_check[sheet0.cell_value(i, course_code).upper()]=1
	if(sheet0.cell_value(i, tag).upper()[0:4]=='H101'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H101[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H101[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H101[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H101[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H101[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H103'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H103[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H103[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H103[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H103[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H103[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H106'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H106[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H106[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H106[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H106[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H106[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H123'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H123[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H123[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H123[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H123[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H123[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H129'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H129[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H129[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H129[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H129[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H129[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H140'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H140[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H140[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H140[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H140[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H140[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H141'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H141[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H141[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H141[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H141[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H141[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H151'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H151[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H151[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H151[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H151[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H151[sheet0.cell_value(i, course_code).upper()]=1
	elif(sheet0.cell_value(i, tag).upper()[0:4]=='H152'):
		if(sheet0.cell_value(i, tag).upper()[4:6]=='CD'):
			course_H152[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H152[sheet0.cell_value(i, course_code).upper()]=1
		else:
			course_H152[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			elec_H152[sheet0.cell_value(i, course_code).upper()]=(sheet0.cell_value(i, course_name))
			unit_H152[sheet0.cell_value(i, course_code).upper()]=1

for i in range(sheet2.ncols):
	if(sheet2.cell_value(0, i).lower()=='id'):
		Id=i
	elif(sheet2.cell_value(0, i).lower()=='name'):
		name=i
	elif(sheet2.cell_value(0, i).lower()=='course'):
		course=i
	elif(sheet2.cell_value(0, i).lower()=='code'):
		code=i

student_H101={}
student_H103={}
student_H106={}
student_H123={}
student_H129={}
student_H140={}
student_H141={}
student_H151={}
student_H152={}
for i in range(1,sheet2.nrows):
	_NAME_=sheet2.cell_value(i, name).upper()
	_ID_=sheet2.cell_value(i, Id).upper()
	_COURSE_=sheet2.cell_value(i, course).upper()
	_CODE_=sheet2.cell_value(i, code).upper()
	if(_ID_[4:8]=='H101'):
		if (_ID_,_NAME_) not in student_H101:
			student_H101[(_ID_,_NAME_)]=list()
			student_H101[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H101[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H103'):
		if (_ID_,_NAME_) not in student_H103:
			student_H103[(_ID_,_NAME_)]=list()
			student_H103[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H103[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H106'):
		print(1)
		if (_ID_,_NAME_) not in student_H106:
			student_H106[(_ID_,_NAME_)]=list()
			student_H106[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H106[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H123'):
		if (_ID_,_NAME_) not in student_H123:
			student_H123[(_ID_,_NAME_)]=list()
			student_H123[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H123[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H129'):
		if (_ID_,_NAME_) not in student_H129:
			student_H129[(_ID_,_NAME_)]=list()
			student_H129[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H129[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H140'):
		if (_ID_,_NAME_) not in student_H140:
			student_H140[(_ID_,_NAME_)]=list()
			student_H140[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H140[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H141'):
		if (_ID_,_NAME_) not in student_H141:
			student_H141[(_ID_,_NAME_)]=list()
			student_H141[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H141[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H151'):
		if (_ID_,_NAME_) not in student_H151:
			student_H151[(_ID_,_NAME_)]=list()
			student_H151[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H151[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
	elif(_ID_[4:8]=='H152'):
		if (_ID_,_NAME_) not in student_H152:
			student_H152[(_ID_,_NAME_)]=list()
			student_H152[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))
		else:
			student_H152[(_ID_,_NAME_)].append((_COURSE_+_CODE_).replace("    "," "))


import xlsxwriter 
workbook = xlsxwriter.Workbook('output.xlsx')
cell_format = workbook.add_format({'bold': True, 'align': 'center'})
cell_format2 = workbook.add_format({'bold': True, 'font_color': 'green', 'align': 'center'})
cell_format3 =workbook.add_format ({'align': 'center'})
for i,j in Dict.items():
	workbook.add_worksheet(j)
for i,j in Dict.items():
	worksheet=workbook.get_worksheet_by_name(j)
	worksheet.write(0,0,'NAME',cell_format)
	worksheet.write(0,1,'ID',cell_format)	
	if(j[0]==' '):
		j=j[1:]
	if(j[-1]==' '):
		j=j[0:len(j)-1]
	col=2
	if(j=='M.E Chemical'):
		length=len(course_H101)
		worksheet.set_column(0, length+6, 30) 
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H101.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Computer Science'):
		length=len(course_H103)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H103.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Mechanical'):
		length=len(course_H106)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H106.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Microelectronics'):
		length=len(course_H123)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H123.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Biotechnology'):
		length=len(course_H129)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H129.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Embedded Systems'):
		length=len(course_H140)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H140.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Design'):
		length=len(course_H141)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H141.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.E Sanitation'):
		length=len(course_H151)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H151.items():
			worksheet.write(0,col,m)
			col=col+1
	elif(j=='M.Phil Liberal Studies'):
		length=len(course_H152)
		worksheet.set_column(0, length+6, 30)
		worksheet.write(0,length+2,'CDC GAINED',cell_format)
		worksheet.write(0,length+3,'ELE GAINED',cell_format)
		worksheet.write(0,length+4,'CDC LEFT',cell_format)
		worksheet.write(0,length+5,'ELE LEFT',cell_format)
		worksheet.write(0,length+6,'PS/ DIST',cell_format)
		worksheet.write(0, length+7, 'BLUE',cell_format)
		for k,m in course_H152.items():
			worksheet.write(0,col,m)
			col=col+1
Y='YES'
N='NO'
for i,j in Dict.items():
	worksheet=workbook.get_worksheet_by_name(j)
	if(i=='H101'):
		row=1
		col=0
		length=len(course_H101)
		for (w,x),z in student_H101.items():
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			dist=0
			for y in z:
				if y in course_H101.keys():
					res = list(course_H101.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H101:
						total_ele+=unit_H101[y]
					elif(y!='BITS G562T' or y!='BITS G563T' or y!='BITS G639' or y!='BITS G629T'):
						total_cdc+=unit_H101[y]
					else:
						if(y=='BITS G562T' or y=='BITS G563T' or y=='BITS G629T'):
							dist=1
						PS=unit_H101[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(6-total_cdc,0),cell_format3)
			if(dist==0):
				worksheet.write(row,length+5,max(7-total_ele,0),cell_format3)
			else:
				worksheet.write(row,length+5,max(3-total_ele,0),cell_format3)

			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)
			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H103'):
		row=1
		col=0
		length=len(course_H103)
		for (w,x),z in student_H103.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			for y in z:
				if y in course_H103.keys():
					res = list(course_H103.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H103.keys():
						total_ele+=unit_H103[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H103[y]
					else:
						PS=unit_H103[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(7-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(6-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)
			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H106'):
		row=1
		col=0
		length=len(course_H106)
		for (w,x),z in student_H106.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			PS=0
			total_ele=0
			blue_count = 0
			for y in z:
				if y in course_H106.keys():
					res = list(course_H106.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H106:
						total_ele+=unit_H106[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H106[y]
					else:
						PS=unit_H106[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(8-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(5-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H123'):
		row=1
		col=0
		length=len(course_H123)
		for (w,x),z in student_H123.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			for y in z:
				if y in course_H123.keys():
					res = list(course_H123.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H101:
						total_ele+=unit_H123[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H123[y]
					else:
						PS=unit_H123[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(7-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(6-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H129'):
		row=1
		col=0
		length=len(course_H129)
		for (w,x),z in student_H129.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			PS=0
			total_ele=0
			blue_count = 0
			for y in z:
				if y in course_H129.keys():
					res = list(course_H129.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H101:
						total_ele+=unit_H129[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H129[y]
					else:
						PS=unit_H129[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(8-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(7-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H140'):
		row=1
		col=0
		length=len(course_H140)
		for (w,x),z in student_H140.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			PS=0
			total_ele=0
			blue_count = 0
			for y in z:
				if y in course_H140.keys():
					res = list(course_H140.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H140:
						total_ele+=unit_H140[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H140[y]
					else:
						PS=unit_H140[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(7-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(6-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H141'):
		row=1
		col=0
		length=len(course_H141)
		for (w,x),z in student_H141.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			for y in z:
					if y in course_H141.keys():
						res = list(course_H141.keys()).index(y)
						worksheet.write(row,res+2,Y,cell_format2)
						if y in blue_check:
							blue_count+=blue_check[y]
						if y in elec_H141:
							total_ele+=unit_H141[y]
						elif(y!='BITS G629T' or y!='BITS G639'):
							total_cdc+=unit_H141[y]
						else:
							PS=unit_H141[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(8-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(5-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H151'):
		row=1
		col=0
		length=len(course_H151)
		for (w,x),z in student_H151.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			for y in z:
				if y in course_H151.keys():
					res = list(course_H151.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H151.keys():
						total_ele+=unit_H151[y]
					elif(y!='BITS G562T' or y!='BITS G563'):
						total_cdc+=unit_H151[y]
					else:
						PS=unit_H151[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(8-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(2-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
	elif(i=='H152'):
		row=1
		col=0
		length=len(course_H152)
		for (w,x),z in student_H152.items():
			worksheet.write(row,0,x,cell_format3)
			worksheet.write(row,1,w,cell_format3)
			total_cdc=0
			total_ele=0
			PS=0
			blue_count = 0
			for y in z:
				if y in course_H152.keys():
					res = list(course_H152.keys()).index(y)
					worksheet.write(row,res+2,Y,cell_format2)
					if y in blue_check:
						blue_count+=blue_check[y]
					if y in elec_H152.keys():
						total_ele+=unit_H152[y]
					elif(y!='BITS G629T' or y!='BITS G639'):
						total_cdc+=unit_H152[y]
					else:
						PS=unit_H152[y]
			worksheet.write(row,length+2,total_cdc,cell_format3)
			worksheet.write(row,length+3,total_ele,cell_format3)
			worksheet.write(row,length+4,max(6-total_cdc,0),cell_format3)
			worksheet.write(row,length+5,max(3-total_ele,0),cell_format3)
			if(PS!=0):
				worksheet.write(row,length+6,PS,cell_format3)

			if(blue_count):
				worksheet.write(row, length+7,blue_count, cell_format3)
			row=row+1
workbook.close() 

import pandas as pd

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


all_dfs = pd.read_excel('output.xlsx', sheet_name=None)
df = pd.concat(all_dfs, ignore_index=True)

df=df[['NAME', 'ID','CDC GAINED','ELE GAINED','CDC LEFT','ELE LEFT','PS/ DIST', 'BLUE']]
append_df_to_excel('output.xlsx', df, sheet_name='Combined', index=False)

print(df)